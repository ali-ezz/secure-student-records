"""
📊 Data Export Utilities
Export data to CSV, Excel, and PDF formats
Respects Flow Control security
"""

import csv
import os
from datetime import datetime
from pathlib import Path


def export_to_csv(data, columns, filename=None, classified=False):
    """
    Export data to CSV file
    
    Args:
        data: List of tuples/lists (table rows)
        columns: List of column names
        filename: Output filename (optional)
        classified: If True, check Flow Control
    
    Returns:
        str: Path to exported file or None if blocked
    """
    if classified:
        # Flow Control: Prevent export of classified data
        return None
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.csv"
    
    # Ensure .csv extension
    if not filename.endswith('.csv'):
        filename += '.csv'
    
    # Export directory
    export_dir = Path.home() / "Downloads" / "SRMS_Exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    
    filepath = export_dir / filename
    
    try:
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(columns)
            
            # Write data
            writer.writerows(data)
        
        return str(filepath)
    
    except Exception as e:
        print(f"Export error: {e}")
        return None


def export_to_excel(data, columns, filename=None, classified=False):
    """
    Export data to Excel file
    
    Requires: openpyxl
    """
    if classified:
        return None
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return export_to_csv(data, columns, filename, classified)  # Fallback to CSV
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.xlsx"
    
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    export_dir = Path.home() / "Downloads" / "SRMS_Exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    filepath = export_dir / filename
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Header styling
        header_fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return str(filepath)
    
    except Exception as e:
        print(f"Excel export error: {e}")
        return None


def export_to_pdf(data, columns, filename=None, title="Data Export", classified=False):
    """
    Export data to PDF file
    
    Requires: reportlab or fpdf2
    """
    if classified:
        return None
    
    try:
        from fpdf import FPDF
    except ImportError:
        print("fpdf2 not installed. Install with: pip install fpdf2")
        return export_to_csv(data, columns, filename, classified)  # Fallback
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{timestamp}.pdf"
    
    if not filename.endswith('.pdf'):
        filename += '.pdf'
    
    export_dir = Path.home() / "Downloads" / "SRMS_Exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    filepath = export_dir / filename
    
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        
        # Title
        pdf.cell(0, 10, title, ln=True, align='C')
        pdf.ln(10)
        
        # Table
        pdf.set_font("Arial", 'B', 10)
        
        # Header
        col_width = 190 / len(columns)  # Distribute evenly
        for col in columns:
            pdf.cell(col_width, 10, str(col)[:20], 1, 0, 'C')
        pdf.ln()
        
        # Data
        pdf.set_font("Arial", '', 9)
        for row in data:
            for value in row:
                pdf.cell(col_width, 8, str(value)[:20], 1, 0, 'C')
            pdf.ln()
        
        pdf.output(str(filepath))
        return str(filepath)
    
    except Exception as e:
        print(f"PDF export error: {e}")
        return None


class ExportDialog:
    """Simple export dialog for choosing format"""
    
    @staticmethod
    def show(parent, data, columns, default_name="export", classified=False):
        """
        Show export dialog
        
        Returns:
            str: Path to exported file or None
        """
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        
        if classified:
            messagebox.showwarning(
                "Flow Control",
                "🔒 Export disabled for classified data.\n\n"
                "Flow Control prevents data from flowing\n"
                "from higher to lower classification levels."
            )
            return None
        
        dialog = tk.Toplevel(parent)
        dialog.title("Export Data")
        dialog.geometry("400x250")
        dialog.resizable(False, False)
        
        # Center
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 400) // 2
        y = (dialog.winfo_screenheight() - 250) // 2
        dialog.geometry(f"+{x}+{y}")
        
        result = [None]
        
        # Header
        ttk.Label(
            dialog,
            text="📊 Export Data",
            font=('Segoe UI', 14, 'bold')
        ).pack(pady=20)
        
        # Format selection
        ttk.Label(dialog, text="Select Format:").pack(anchor='w', padx=40)
        
        format_var = tk.StringVar(value="CSV")
        format_frame = ttk.Frame(dialog)
        format_frame.pack(pady=10)
        
        for fmt in ['CSV', 'Excel', 'PDF']:
            ttk.Radiobutton(
                format_frame,
                text=fmt,
                variable=format_var,
                value=fmt
            ).pack(side='left', padx=10)
        
        # Filename
        ttk.Label(dialog, text="Filename:").pack(anchor='w', padx=40, pady=(10, 0))
        filename_var = tk.StringVar(value=default_name)
        ttk.Entry(dialog, textvariable=filename_var, width=40).pack(padx=40, pady=5)
        
        def do_export():
            fmt = format_var.get()
            filename = filename_var.get()
            
            if fmt == 'CSV':
                path = export_to_csv(data, columns, filename)
            elif fmt == 'Excel':
                path = export_to_excel(data, columns, filename)
            else:  # PDF
                path = export_to_pdf(data, columns, filename)
            
            if path:
                messagebox.showinfo(
                    "Success",
                    f"✅ Data exported successfully!\n\nSaved to:\n{path}"
                )
                result[0] = path
                dialog.destroy()
            else:
                messagebox.showerror(
                    "Error",
                    "❌ Failed to export data.\nCheck console for details."
                )
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="Export", command=do_export).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side='left', padx=5)
        
        dialog.wait_window()
        return result[0]


if __name__ == "__main__":
    # Test
    test_data = [
        (1, 'John Doe', 'john@example.com', 'Admin'),
        (2, 'Jane Smith', 'jane@example.com', 'Student'),
        (3, 'Bob Johnson', 'bob@example.com', 'Instructor'),
    ]
    
    test_columns = ['ID', 'Name', 'Email', 'Role']
    
    print("Testing CSV export...")
    csv_path = export_to_csv(test_data, test_columns, "test_export")
    print(f"CSV exported to: {csv_path}")
    
    print("\nTesting Excel export...")
    excel_path = export_to_excel(test_data, test_columns, "test_export")
    print(f"Excel exported to: {excel_path}")
